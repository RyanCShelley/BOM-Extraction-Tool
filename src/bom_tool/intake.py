from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .models import IntakeField


def _to_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_intake_fields(intake_path: Path) -> list[IntakeField]:
    wb = load_workbook(filename=intake_path, data_only=True)
    if "IntakeForm BOM" not in wb.sheetnames:
        raise ValueError("Expected sheet 'IntakeForm BOM' in intake workbook.")

    ws = wb["IntakeForm BOM"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    fields: list[IntakeField] = []

    for row in rows:
        section = _to_str(row[0] if len(row) > 0 else "")
        field_number = _to_str(row[1] if len(row) > 1 else "")
        field_name = _to_str(row[2] if len(row) > 2 else "")
        field_type = _to_str(row[3] if len(row) > 3 else "")
        allowed_values = _to_str(row[4] if len(row) > 4 else "")
        source_process = _to_str(row[5] if len(row) > 5 else "")
        data_point_location = _to_str(row[6] if len(row) > 6 else "")

        if not section or not field_name:
            continue

        fields.append(
            IntakeField(
                section=section,
                field_number=field_number,
                field_name=field_name,
                field_type=field_type,
                allowed_values=allowed_values,
                source_process=source_process,
                data_point_location=data_point_location,
            )
        )

    return fields
